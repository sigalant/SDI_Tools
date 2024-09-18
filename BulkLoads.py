#Imports
import os
import tkinter as tk
from tkinter import filedialog
import re
import sys
import traceback
from datetime import date

from PIL import Image, ImageTk
import openpyxl as opx

import LogErrors
import FindHeaders

def resource_path(rel_path):
    try:
        base_path = sys.MEIPASS
    except Exception:
        base_path = os.path.abspath("./_internal")
    return os.path.join(base_path, rel_path)

#Holds I/O filepaths
inputFilepath = ""
outputFilepath = ""


#===== Root window =============================================================================================================
root = tk.Tk()

root.title("SDI Bulk Loads Formatting Tool")
root.geometry("800x450")

menubar = tk.Menu(root)
helpMenu = tk.Menu(menubar, tearoff=0)
helpMenu.add_command(label="Help", command = lambda:os.startfile(resource_path('Help.html')))
helpMenu.add_command(label="Examples", command = lambda:os.startfile(filedialog.askopenfilename(initialdir=resource_path('./Example Files'))))
menubar.add_cascade(label="Help", menu=helpMenu)
root.config(menu=menubar)

#Opens popup window with traceback if unhandled exception
def handle_exception(exc,val,tb):
    top = tk.Toplevel(root)
    top.geometry("800x400")
    top.title("ERROR")
    tk.Label(top, text="ERROR: \n", font=(25)).pack(pady=10)
    tk.Label(top, text="".join(traceback.format_exception(exc,val,tb))).pack()
    tk.Button(top, text="OK", command = top.destroy).pack()
    root.wait_window(top)
    LogErrors.handle_exception(exc,val,tb)

root.report_callback_exception = handle_exception

ico= Image.open(resource_path("SDI_Logo.ico"))
photo = ImageTk.PhotoImage(ico)
root.wm_iconphoto(False, photo)

#For displaying errors to user
errorFrame = tk.Frame(root)
errorFrame.pack(side=tk.BOTTOM)
errorMsg = tk.Label(errorFrame, text="")
errorMsg.pack(pady=40)
#===============================================================================================================================

def formatFile(voltList):
    #============== Check Nothing is Missing ============================================================================================================
    #Stop if I/O paths not provided
    if inputFilepath == '':
        errorMsg.config(text= "Error: No input file selected")
        return
    if outputFilepath == '':
        errorMsg.config(text="Error: No output folder selected")
        return
    #message displayed if everything works
    errorMsg.config(text="File Successfully Formatted")

    wb = opx.load_workbook(inputFilepath)
    sheet = wb.active

    temp = inputFilepath.split('/')
    filename = temp[(len(temp)-1)].split('.')[0] + "_formatted.xlsx"
    newFile = outputFilepath+"/"+filename
    wbNew=opx.Workbook()
    sheetNew=wbNew.active

    hDict = FindHeaders.FindHeaders(inputFilepath)#Holds index of important values

    inList = ['amps', 'kw', 'gph', 'btus', 'exh cfm', 'supply cfm', 'volts', 'ph', 'heat rejection', 'no', 'qty']
    metricInList = ['amps', 'volts','ph','lph', 'gas kw', 'exh (m^3/h)', 'supply (m^3/h)', 'heat rejection watts','no','qty']
    summingIndexes = ['kw', 'gph', 'btus', 'exh cfm', 'supply cfm', 'heat rejection']

    missing = []
    for header in inList:
        if header not in hDict:
            missing.append(header)
            
    metricMissing = []
    for header in metricInList:
        if header not in hDict:
            metricMissing.append(header)

    if missing:
        if metricMissing:
            errorStr = ''
            if len(metricMissing) < len(missing):
                errorStr = "ERROR: The following header(s) was not found " + ', '.join(missing)
            else:
                errorStr = "ERROR: The following header(s) was not found " + ', '.join(metricMissing)
            errorMsg.config(text=errorStr)
            return
        inList = metricInList
        summingIndexes = ['amps','lph','gas kw', 'exh (m^3/h)', 'supply (m^3/h)', 'heat rejection watts']
    
    sheetData = []#Holds all data from input sheet
    

    #================= Copy all data to a 2D-List =========================================================================================================
    for row in sheet.rows:
        rowData = ['' for i in range(len(hDict))]#Holds all data from a row
        mult=1.0 #Holds value in paranthesis from amp cell
        

        for head in hDict:
            cellData = row[hDict[head]].value
            
            #Remove commas for better number processing
            if type(cellData) == str:
                cellData = cellData.replace(',','')

            # Adjust for 'newlines' in cell
            if "_x000D_" in str(cellData) and (head in ['volts', 'ph', 'amps'] or head in summingIndexes):
                cellData = cellData.split("_x000D_")
                for i in range(len(cellData)):
                    if '(' in cellData[i] and head not in ['volts','ph']:
                        fList = [float(t) for t in re.findall(r'\d+\.?\d*', cellData[i])]
                        cellData[i] = fList[0]*fList[1]
                    else:
                        cellData[i] = [float(t) for t in re.findall(r'\d+\.?\d*', cellData[i])][0]
                if head not in ['volts', 'ph', 'amps']: # maybe also kw?
                    cellData = sum(cellData)
            
            #Adjust for quantity x: (x)...A
            elif ')' in str(cellData) and (head in summingIndexes or head == 'amps') and row[hDict[head]].row>1:#head == 'amps':
                cellData = [float(s) for s in re.findall(r'\d+\.?\d*',cellData)]
                if head == 'amps': mult= cellData[0]
                cellData = cellData[0] * cellData[1]

            #Strip summing fields of everything except a number
            if type(cellData) == str and (head in summingIndexes or head in ['volts','ph','amps']) and row[hDict[head]].row>1:
                if head == 'volts' and '/' in cellData:
                    cellData = '208'
                cellData = [float(t) for t in re.findall(r'\d+\.?\d*',cellData)]
                if len(cellData) > 1:
                    print(str(cellData) + ":" + str(head) + ":" + str(row[hDict[head]].row))
                else:
                    cellData = float(cellData[0])
                        
            rowData[hDict[head]] = cellData
        
        #If an electrical column had a 'newline'(_x000D_), then give each value its own row (only works for 2 values)
        if type(rowData[hDict['volts']]) == list or type(rowData[hDict['ph']]) == list or type(rowData[hDict['amps']]) == list:
            vs= rowData[hDict['volts']] if isinstance(rowData[hDict['volts']],list) else [rowData[hDict['volts']]]
            ps= rowData[hDict['ph']] if isinstance(rowData[hDict['ph']],list) else [rowData[hDict['ph']]]
            ams= rowData[hDict['amps']] if isinstance(rowData[hDict['amps']],list) else [rowData[hDict['amps']]]
            
            for i in range(len(max([vs,ps,ams], default=(), key=len))):
                newRow = rowData.copy()
                if i<len(vs):
                    newRow[hDict['volts']] = vs[i]
                else:
                    newRow[hDict['volts']] = vs[0]
                if i<len(ps):
                    newRow[hDict['ph']] = ps[i]
                else:
                    newRow[hDict['ph']] = ps[0]
                if i<len(ams):
                    newRow[hDict['amps']] = ams[i]/mult
                else:
                    newRow[hDict['amps']] = ams[0]/mult
                if(i):
                    for entry in summingIndexes:
                        if entry not in ('amps','kw'):
                            newRow[hDict[entry]] = None
                sheetData.append(newRow)
        else:
            sheetData.append(rowData)
            
    

#============ Paste data to new sheet ================================================================================================================================    
    #Excel File Heading
    sheetNew.row_dimensions[1].height=30
    sheetNew.row_dimensions[2].height=30.75
    sheetNew.row_dimensions[3].height=23.25
    sheetNew.row_dimensions[4].height=27
    sheetNew.row_dimensions[5].height=30
    sheetNew.row_dimensions[6].height=15.75
    sheetNew.row_dimensions[7].height=5

    sheetNew.column_dimensions['A'].width = 10
    sheetNew.column_dimensions['B'].width = 5
    sheetNew.column_dimensions['C'].width = 40.1
    sheetNew.column_dimensions['D'].width = 10
    sheetNew.column_dimensions['E'].width = 10
    sheetNew.column_dimensions['F'].width = 10
    sheetNew.column_dimensions['G'].width = 10
    sheetNew.column_dimensions['H'].width = 10
    if 'remarks' in hDict:
        sheetNew.column_dimensions[chr(65 + hDict['remarks'])].width = 30

    sheetNew.page_setup.paperSize = sheetNew.PAPERSIZE_TABLOID
    sheetNew.sheet_properties.pageSetUpPr.fitToPage = True
    sheetNew.page_setup.fitToHeight = False
    sheetNew.page_setup.orientation = sheetNew.ORIENTATION_LANDSCAPE
    
    img = opx.drawing.image.Image(resource_path("SDI_Logo.PNG"))
    img.height=40
    img.width=65
    sheetNew.add_image(img, "A1")

    sheetNew['A3'] = "________ Preliminary Utility Schedule"
    sheetNew['A3'].font = opx.styles.Font(size=24, bold=True)
    sheetNew['A4'] = str(date.today().strftime("%B %d, %Y"))
    sheetNew['A4'].font = opx.styles.Font(size=18, bold=True)
    sheetNew.append([''])        

    sheetNew.append(sheetData[0])
    sheetNew.freeze_panes = sheetNew['C7']

    #Add data to sheet
    for row in range(1,len(sheetData)):

        #Check if area header
        if sheetData[row][hDict['no']] != None and sheetData[row][hDict['no']] == str(sheetData[row][hDict['no']]).upper() and sheetData[row][hDict['qty']] == None:
            sheetNew.append([sheetData[row][hDict['no']]])
            sheetNew[sheetNew.max_row][hDict['no']].font = opx.styles.Font(size=14, color='FFFFFF', bold=True)
            for i in range(sheetNew.max_column):
                sheetNew[sheetNew.max_row][i].fill = opx.styles.PatternFill(fgColor="002060", fill_type="solid")
            continue
        
        
        #Add data
        if sheetData[row][hDict['volts']] != None and 'kw' in hDict:
            if 'amps' in hDict:
                sheetData[row][hDict['kw']] = "=IF("+chr((hDict['kw']-2)+65)+str(row+7)+">1,(1.732*"+chr((hDict['kw']-3)+65)+str(row+7)+"*"+chr((hDict['kw']-1)+65)+str(row+7)+")/1000,("+chr((hDict['kw']-3)+65)+str(row+7)+"*"+chr((hDict['kw']-1)+65)+str(row+7)+")/1000)"
            else:
                sheetData[row][hDict['kw']] = [float(s) for s in re.findall(r'\d+\.?\d*',sheetData[row][indexDict['kw']])][0]
                
        sheetNew.append(sheetData[row])
        
        #Alignments/Formatting
        sheetNew[sheetNew.max_row][hDict['remarks']].alignment = opx.styles.Alignment(wrap_text=True)
        
        sheetNew[sheetNew.max_row][0].alignment = opx.styles.Alignment(horizontal='right')
        sheetNew[sheetNew.max_row][1].alignment = opx.styles.Alignment(horizontal='center')
        
        if sheetData[row][hDict['volts']] != None and float(sheetData[row][hDict['volts']]) not in voltList:
            sheetNew[row+7][hDict['volts']].fill = opx.styles.PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')

        if 'amps' in hDict:
            sheetNew[sheetNew.max_row][hDict['amps']].number_format="0.0"
        if 'kw' in hDict:
            sheetNew[sheetNew.max_row][hDict['kw']].number_format="0.00"

    #Sums
    sheetNew.append([""])
    sheetNew.append(["Total"])
    sheetNew[sheetNew.max_row][0].font = opx.styles.Font(bold=True)
    
    for header in summingIndexes:
        sheetNew[sheetNew.max_row][hDict[header]].value = "=SUM("+chr(hDict[header]+65)+str(7)+":"+chr(hDict[header]+65)+str(sheetNew.max_row-1)+")"
        sheetNew[sheetNew.max_row][hDict[header]].number_format="#,##0"
        sheetNew[sheetNew.max_row][hDict[header]].font = opx.styles.Font(bold=True)
        
    wbNew.save(newFile)


frame = tk.Frame(root)
frame.pack(padx=40, pady=20)

voltFrame = tk.Frame(root)
voltFrame.pack(pady=20)

fileFrame = tk.Frame(root)
fileFrame.pack()

bottomFrame = tk.Frame(root)
bottomFrame.pack()

in_text = tk.Label(fileFrame, text="The input file is: " + inputFilepath)
in_text.pack(side=tk.TOP)

out_text = tk.Label(fileFrame, text = "The output folder is: " + outputFilepath)
out_text.pack(side=tk.TOP)

voltLabel = tk.Label(voltFrame, text="Enter acceptable voltages:")
voltLabel.pack(side=tk.TOP)
voltEntries = []

#Switch to next text box when 'TAB' or 'RETURN' key pressed
def tab_pressed(event):
    if voltEntries.index(event.widget)+1<len(voltEntries):
        voltEntries[voltEntries.index(event.widget)+1].focus_set()
    else:
        voltEntries[0].focus_set()
    return "break" 

#Add a new text box when the '+' button is pressed
def addVolt(butts):
    voltText = tk.Text(voltFrame, height = 1, width = 10)
    voltText.bind('<Tab>', tab_pressed)
    voltText.bind('<Return>', tab_pressed)
    voltText.pack(padx=10, pady=10, side = tk.LEFT)
    voltEntries.append(voltText)
    if(len(voltEntries) >= 5):
        butts[0].pack_forget()
    butts[1].pack(side=tk.RIGHT)

#Delete a text box when the '-' button is pressed
def removeVolt(butts):
    voltText = voltEntries.pop()
    voltText.destroy()
    if(len(voltEntries) <= 1):
        butts[1].pack_forget()
    butts[0].pack(side=tk.RIGHT)

voltButtons=[None,None]
addVoltButton = tk.Button(voltFrame, text = "+", command=lambda:addVolt(voltButtons))
removeVoltButton = tk.Button(voltFrame, text = "-", command=lambda:removeVolt(voltButtons))
voltButtons[0]=addVoltButton
voltButtons[1]=removeVoltButton



for i in range(3):
    voltText = tk.Text(voltFrame, height = 1, width = 10)
    voltText.bind('<Tab>', tab_pressed)
    voltText.bind('<Return>', tab_pressed)
    match i:
        case 0:
            voltText.insert(tk.END, "120")
        case 1:
            voltText.insert(tk.END, "208")
        case 2:
            voltText.insert(tk.END, "480")
    voltText.pack(padx=10, pady=10, side = tk.LEFT)
    voltEntries.append(voltText)

removeVoltButton.pack(padx=1,pady=1,side=tk.RIGHT)
addVoltButton.pack(padx=1,pady=1, side =tk.RIGHT)


def form():
    try:
        voltList = [float(t.get("1.0", 'end-1c')) for t in voltEntries]
    except:
        errorMsg.config(text="Error: All acceptable voltages should be numbers")
        return

    formatFile(voltList)

format_button = tk.Button(bottomFrame, text="format file", command=form)
format_button.pack(padx=10, pady=10, side=tk.BOTTOM)

def getFilepath():
    global inputFilepath
    inputFilepath = filedialog.askopenfilename(filetypes=(("Microsoft Excel Worksheet", "*.xlsx"),))
    in_text.config(text="The input file is: " + inputFilepath)

def getOutputFolder():
    global outputFilepath
    outputFilepath = filedialog.askdirectory()
    out_text.config(text="The output folder is: " + outputFilepath)

in_file = tk.Button(frame, text="select input file", command=getFilepath)
in_file.pack(padx=20, pady=15, side=tk.LEFT)

out_folder = tk.Button(frame, text="select output folder", command=getOutputFolder)
out_folder.pack(padx=20, pady=15, side=tk.LEFT)

root.mainloop()
    
