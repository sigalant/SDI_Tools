#Imports

import tkinter as tk
import openpyxl as opx
import tkinter.ttk as ttk
from datetime import date
from tkinter import filedialog
from PIL import Image, ImageTk
import sys
import traceback
import LogErrors
import os
import FindHeaders


#File Paths
inputFilepath = ""
outputFilepath = ""

#Default percent markup
markup = 18

#============ Tkinter GUI ======================================================================================================
root = tk.Tk()

root.title("SDI Budget Formatting Tool")
root.geometry("800x500")

ico = Image.open("V:\\Budget\\AutoQuotes Budget Script\\SDI Logo.jpg")
photo = ImageTk.PhotoImage(ico)
root.wm_iconphoto(False, photo)

#Help Menu
menubar = tk.Menu(root)
helpMenu = tk.Menu(menubar, tearoff=0)
helpMenu.add_command(label="Help", command=lambda:os.startfile('Help.html'))
helpMenu.add_command(label="Examples", command = lambda:os.startfile(filedialog.askopenfilename(initialdir="./Example Files")))
menubar.add_cascade(label="Help", menu=helpMenu)
root.config(menu=menubar)

#Opens popup window with traceback of unhandled exceptions in tkinter window  
def handle_exception(exc,val,tb):
    top=tk.Toplevel(root)
    top.geometry("800x400")
    top.title("ERROR")
    tk.Label(top,text="ERROR: \n", font=(25)).pack(pady=10)
    tk.Label(top,text="".join(traceback.format_exception(exc,val,tb))).pack()
    tk.Button(top, text="OK", command = top.destroy).pack()
    root.wait_window(top)
    LogErrors.handle_exception(exc,val,tb)

root.report_callback_exception = handle_exception

#GUI widgets
inputFrame = tk.Frame(root)
inputFrame.pack()

markupFrame = tk.Frame(inputFrame)
markupFrame.pack(side=tk.LEFT)

markupLabel = tk.Label(markupFrame, text= "Markup Percent")
markupLabel.pack()

spinbox = ttk.Spinbox(markupFrame, from_=0, to=100, format='%10.2f %%')
spinbox.set('{:10.2f} %'.format(18))
spinbox.pack()

goodThroughFrame = tk.Frame(inputFrame)
goodThroughFrame.pack(side = tk.LEFT, padx=20,pady=40)

gtLabel = tk.Label(goodThroughFrame, text="Enter Good Through Date: ")
gtLabel.pack()

gtTextBox = tk.Entry(goodThroughFrame)
gtTextBox.pack()

#Error Message Label
errorFrame = tk.Frame(root)
errorFrame.pack(side=tk.BOTTOM)
errorMsg = tk.Label(errorFrame, text="")
errorMsg.pack(pady = 50)

#=====================================================================================================================================

#Take Autoquotes export, and format it
def formatFile():

#====== CHECK FOR FILEPATHS AND HEADERS ====================================================================================================================
    
    if inputFilepath == "":
        errorMsg.config(text="Error: No input file selected")
        return
    
    if outputFilepath == "":
        errorMsg.config(text="Error: No output folder selected")
        return

    #Message displayed if function completes without errors
    errorMsg.config(text="Spreadsheet Successfully Formatted")

    #Open Excel File 
    wb = opx.load_workbook(inputFilepath)
    sheet = wb.active
    
    #Get all headers
    hDict = FindHeaders.FindHeaders(inputFilepath)
    headers = ['itemno','qty','category','sell','remarks','spec','model','unit','selltotal'] #Headers used in this function
    missing = [] #Holds missing headers
    #Check for missing headers
    for head in headers:
        if head not in hDict:
            missing.append(head)
    #Display missing headers
    if missing:
        errorMsg.config(text="Warning: The following header(s) may be missing: " + ", ".join(missing))


#========== COPY DATA FROM EXCEL SHEET ======================================================================================================================
    
    #Fill 2D Array with information
    data = [] #To hold all project data for each item

    for row in sheet.rows:
        #Collect data from each row (Skipping empty cells)

        rowData = ["" for i in range(len(hDict))] #To hold data for a single item
        
        #row[0].fill.start_color.index[:2]   # Alpha (unimportant)
        r = int(row[0].fill.start_color.index[2:4], 16)  # Red
        g = int(row[0].fill.start_color.index[4:6], 16)  # Green
        b = int(row[0].fill.start_color.index[6:8], 16)  # Blue
        

        #Skip first row and non-white rows
        if (('itemno' in hDict and str(row[hDict['itemno']].value).lower() == "ItemNo".lower()) 
            or (r < int('ef',16) or g < int('ef', 16) or b < int('ef', 16))):
            continue
        
        for header in hDict:
            #Add each header value for each item. If header not found, add empty value
            try:
                assert (hDict[header] != -1)
                rowData[hDict[header]] = row[hDict[header]].value
            except Exception:
                rowData[hDict[header]] = ''

        #Remove Cost and Qty of SpareNo items
        if 'category' in hDict and rowData[hDict['category']] != None: 
            rowData[hDict['category']] = str(rowData[hDict['category']]).upper()
            if rowData[hDict['category']] == "SPARENO" and 'qty' in hDict and 'sell' in hDict:
                rowData[hDict['category']] = 'SpareNo'
                rowData[hDict['qty']] = '-'
                rowData[hDict['sell']] = ''

        #Check if item not in contract
        if 'sell' in hDict and 'remarks' in hDict and type(rowData[hDict['remarks']]) == str:
            if ("os&e" in str(rowData[hDict['remarks']]).lower() or "by vendor" in str(rowData[hDict['remarks']]).lower() or 'contractor' in str(rowData[hDict['remarks']]).lower() or 'millwork' in str(rowData[hDict['remarks']]).lower()):
                rowData[hDict['sell']] = 'NIC'
            elif 'model' in hDict and "os&e" in str(rowData[hDict['model']]).lower():
                rowData[hDict['sell']] = 'NIC'
        #Add data for row to 2D array of all project data
        data.append(rowData)

#========== HEADER INFORMATION ==================================================================================================================================================================================================================================================================================        

    #Create new Excel File
    
    temp = inputFilepath.split('/')

    filename = temp[(len(temp)-1)].split('.')[0] + "_formatted.xlsx"

    newFile = outputFilepath + "/" + filename

    wbNew = opx.Workbook()


    #Fill information
    
    sheetNew = wbNew.active

    sheetNew.sheet_view.showGridLines= False

    sheetNew.row_dimensions[1].height = 30
    sheetNew.row_dimensions[2].height = 30.75
    sheetNew.row_dimensions[3].height = 23.25
    sheetNew.row_dimensions[4].height = 27
    sheetNew.row_dimensions[5].height = 30
    sheetNew.row_dimensions[6].height = 15.75
    sheetNew.row_dimensions[7].height = 18.75

    sheetNew.column_dimensions['A'].width = 10
    sheetNew.column_dimensions['B'].width = 13.3
    sheetNew.column_dimensions['C'].width = 30.1
    sheetNew.column_dimensions['D'].width = 28.9
    sheetNew.column_dimensions['E'].width = 15.9
    sheetNew.column_dimensions['F'].width = 19
    sheetNew.column_dimensions['G'].width = 3.7
    sheetNew.column_dimensions['H'].width = 64.9


    headerBorder = opx.styles.borders.Border(top=opx.styles.borders.Side(style='thick', color='80002060'), bottom=opx.styles.borders.Side(style='thick'))
    
    img = opx.drawing.image.Image("V:\\Budget\\AutoQuotes Budget Script\\SDI Logo.jpg")
    img.height = 40
    img.width = 65
    sheetNew.add_image(img, 'A1')
    

    sheetNew['A2'] = ""
    sheetNew['A2'].font = opx.styles.Font(size=24, bold=True)

    sheetNew['A3'] = str(date.today().strftime("%B %d, %Y"))
    sheetNew['A3'].font = opx.styles.Font(size=18, bold=True) 

    sheetNew['A5'] = "Item"
    sheetNew['A5'].border = headerBorder
    sheetNew['A5'].alignment = opx.styles.Alignment(horizontal = 'left', vertical = 'center')

    sheetNew['B5'] = "Qty"
    sheetNew['B5'].border = headerBorder
    sheetNew['B5'].alignment = opx.styles.Alignment(horizontal = 'center', vertical = 'center')

    sheetNew['C5'] = "Description"
    sheetNew['C5'].border = headerBorder
    sheetNew['C5'].alignment = opx.styles.Alignment(horizontal = 'left', vertical = 'center')

    sheetNew['D5'] = "Model"
    sheetNew['D5'].border = headerBorder
    sheetNew['D5'].alignment = opx.styles.Alignment(horizontal = 'left', vertical = 'center')

    sheetNew['E5'] = "Unit Cost"
    sheetNew['E5'].border = headerBorder
    sheetNew['E5'].alignment = opx.styles.Alignment(horizontal = 'right', vertical = 'center')

    sheetNew['F5'] = "Total"
    sheetNew['F5'].border = headerBorder
    sheetNew['F5'].alignment = opx.styles.Alignment(horizontal = 'right', vertical = 'center')

    sheetNew['H5'] = "Remarks"
    sheetNew['H5'].border = headerBorder
    sheetNew['H5'].alignment = opx.styles.Alignment(horizontal = 'left', vertical = 'center')

    rowNum = 6
    
    errorShown = False
    
#============= PASTE COPIED DATA =======================================================================================================================================================================================================================================================

    #Copy previously collected data to a new XL sheet
    for i in range(len(data)):
        
        if 'spec' in hDict and 'itemno' in hDict and data[i][hDict['spec']] != None and data[i][hDict['spec']] == str(data[i][hDict['spec']]).upper() and data[i][hDict['itemno']] == None:
            
            #If there isn't title, use the first location header as a title
            if sheetNew['A2'].value == "":
                sheetNew['A2'] = data[i][hDict['spec']]
                continue
            rowNum = rowNum+1
            c = "A"+str(rowNum)
            sheetNew[c] = data[i][hDict['spec']]
            sheetNew[c].font = opx.styles.Font(size = 14, color = 'FFFFFF', bold = True)
            sheetNew[c].fill = opx.styles.PatternFill(fgColor="002060", fill_type="solid")
            sheetNew["B"+str(rowNum)].fill = opx.styles.PatternFill(fgColor = "002060", fill_type="solid")
           
            sheetNew["C"+str(rowNum)].fill = opx.styles.PatternFill(fgColor = "002060", fill_type="solid")
            
            sheetNew["D"+str(rowNum)].fill = opx.styles.PatternFill(fgColor = "002060", fill_type="solid")
            sheetNew["E"+str(rowNum)].fill = opx.styles.PatternFill(fgColor = "002060", fill_type="solid")
            sheetNew["F"+str(rowNum)].fill = opx.styles.PatternFill(fgColor = "002060", fill_type="solid")
            
            sheetNew["H"+str(rowNum)].fill = opx.styles.PatternFill(fgColor = "002060", fill_type="solid")

            
            rowNum = rowNum+1
        else:
            if 'itemno' in hDict and data[i][hDict['itemno']] == None and not errorShown:
                tk.messagebox.showerror('Formatting Error', "Error: Please collapse all items in autocad before export")
                errorShown = True

            if 'itemno' in hDict: sheetNew[("A"+str(rowNum))] = data[i][hDict['itemno']]
            if 'unit' in hDict and 'qty' in hDict and data[i][hDict['unit']] != None and str(data[i][hDict['unit']]).lower() == 'ft':
                sheetNew[("B"+str(rowNum))] = 1
            else:
                if 'qty' in hDict: sheetNew[("B"+str(rowNum))] = data[i][hDict['qty']]
            if 'category' in hDict: sheetNew[("C"+str(rowNum))] = data[i][hDict['category']]
            sheetNew['C'+str(rowNum)].alignment = opx.styles.Alignment(wrap_text = True)
            if 'model' in hDict: sheetNew[("D"+str(rowNum))] = data[i][hDict['model']]
            if 'sell' in hDict: sheetNew[("E"+str(rowNum))] = data[i][hDict['sell']]
            sheetNew[("E"+str(rowNum))].number_format = "$#,##0.00"
            try:
                sheetNew[("F"+str(rowNum))] = float(data[i][hDict['sell']])*float(data[i][hDict['qty']])
            except Exception:
                if 'sell' in hDict: sheetNew[("F"+str(rowNum))] = data[i][hDict['sell']]
            if ('unit' in hDict and data[i][hDict['unit']] != None and 'ft' in str(data[i][hDict['unit']]).lower()) or ('remarks' in hDict and data[i][hDict['remarks']] != None and 'custom fab' in str(data[i][hDict['remarks']]).lower()):
                if 'selltotal' in hDict: sheetNew['F'+str(rowNum)] = data[i][hDict['selltotal']]
            sheetNew["F"+str(rowNum)].number_format = "$#,##0.00"
            if 'remarks' in hDict: sheetNew[("H"+str(rowNum))] = data[i][hDict['remarks']]
            sheetNew['H'+str(rowNum)].alignment = opx.styles.Alignment(wrap_text=True)
            sheetNew[("H"+str(rowNum))].font = opx.styles.Font(color = '595959')
            sheetNew['E'+str(rowNum)].alignment = opx.styles.Alignment(horizontal = 'right', vertical = 'center')
            sheetNew['F'+str(rowNum)].alignment = opx.styles.Alignment(horizontal = 'right', vertical = 'center')
        rowNum = rowNum+1

#============= ADD FOOTER INFORMATION ============================================================================================================================================================================================================================================================

    for i in range(6):
        sheetNew[rowNum][i].border = opx.styles.borders.Border(bottom=opx.styles.borders.Side(style='thick'))

    sheetNew[("E" + str(rowNum+1))] = "Equipment SubTotal"
    sheetNew['E'+str(rowNum+1)].alignment = opx.styles.Alignment(horizontal = 'right', vertical = 'center')

    sheetNew[("E" + str(rowNum+2))] = "Delivery, Installation, Set-in Place"
    sheetNew['E'+str(rowNum+2)].alignment = opx.styles.Alignment(horizontal = 'right', vertical = 'center')

    sheetNew[("E" + str(rowNum+3))] = "Total"
    sheetNew['E'+str(rowNum+3)].alignment = opx.styles.Alignment(horizontal = 'right', vertical = 'center')
    sheetNew['E'+str(rowNum+3)].font = opx.styles.Font(bold=True)

    sheetNew[("F" + str(rowNum+1))] = "=SUM(F6:F" + str(rowNum) + ")"
    sheetNew['F'+str(rowNum+1)].number_format = '$#,##0.00'
    try:
        sheetNew[("F" + str(rowNum+2))] = "=F" + str(rowNum+1) +"*"+ str(float(spinbox.get().split()[0])/100.0)
    except:
        sheetNew[("F"+str(rowNum+2))] = "=F"+str(rowNum+1) +"*"+str(float(18.0/100.0))
    sheetNew['F'+str(rowNum+2)].number_format = '$#,##0.00'
    sheetNew[("F" + str(rowNum+3))] = "=SUM(F" + str(rowNum+1) + ":F" + str(rowNum+2) + ")"
    sheetNew['F'+str(rowNum+3)].number_format = '$#,##0.00'
    sheetNew['F'+str(rowNum+3)].font = opx.styles.Font(bold=True)

    for i in range(3,6):
        sheetNew[rowNum+2][i].border = opx.styles.borders.Border(bottom=opx.styles.borders.Side(style='thick'))

    sheetNew.merge_cells(start_row=rowNum+6, start_column=1, end_row=rowNum+6, end_column=3)
    sheetNew['A'+str(rowNum+6)] = "QUALIFICATIONS"
    sheetNew['A'+str(rowNum+6)].font = opx.styles.Font(bold=True, color='FFFFFF')
    sheetNew['A'+str(rowNum+6)].fill = opx.styles.PatternFill(fgColor = '002060', fill_type='solid')

    sheetNew['A'+str(rowNum+7)] = "1. Price does not include project contingency."
    sheetNew['A'+str(rowNum+7)].font = opx.styles.Font(color ='595959')

    sheetNew['A'+str(rowNum+8)] = "2. Price does not include sales tax or use taxes."
    sheetNew['A'+str(rowNum+8)].font = opx.styles.Font(color ='595959')

    if gtTextBox.get() != "":
        sheetNew['A'+str(rowNum+9)] = ("3. Price is good through " + gtTextBox.get())
    else:
        sheetNew['A'+str(rowNum+9)] = "3. Price is good through _____________"
    sheetNew['A'+str(rowNum+9)].font = opx.styles.Font(color ='595959')
#=======================================================================================================================================================================================================================================================================================================
    wbNew.save(newFile)





# More tkinter stuff
frame = tk.Frame(root)
frame.pack(padx=40, pady=40)


fileFrame = tk.Frame(root)
fileFrame.pack()

bottomFrame = tk.Frame(root)
bottomFrame.pack()

in_text = tk.Label(fileFrame, text="The input file is: " + inputFilepath)
in_text.pack(side=tk.TOP)

out_text = tk.Label(fileFrame, text="The output folder is: " + outputFilepath)
out_text.pack(side=tk.TOP)

format_button = tk.Button(bottomFrame, text="format file", command=formatFile)
format_button.pack(padx=10, pady=20, side=tk.BOTTOM)


def getFilepath():
    global inputFilepath
    inputFilepath = filedialog.askopenfilename(filetypes = (("Microsoft Excel Worksheet", "*.xlsx"),))
    in_text.config(text= "The input file is: " + inputFilepath)

def getOutputFolder():
    global outputFilepath 
    outputFilepath = filedialog.askdirectory()
    out_text.config(text= "The output folder is: " + outputFilepath)

in_file = tk.Button(frame, text="select input file", command=getFilepath)
in_file.pack(padx=10, pady=15, side=tk.LEFT)

out_folder = tk.Button(frame, text="select output folder", command=getOutputFolder)
out_folder.pack(padx=10, pady=15, side=tk.LEFT)


root.mainloop()
